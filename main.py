"""
ES Windows Order Form Automation — Desktop UI
Launches a Tkinter GUI for selecting an Excel file, entering the order number,
and automating form entry via Selenium.

Usage:
    python3 main.py
"""

import os
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

from form_filler import FormFiller, FormFillerError
import form_selectors as sel

CHROME_PROFILE_DIR = os.path.join(os.path.expanduser("~"), ".esw-automation-profile")
VALID_PRODUCT_TYPES = {"WINDOW", "DOOR", "STOREFRONT", "SHAPE", "MULLION"}


def read_excel(file_path):
    """Read line items from the 'Data Entry' sheet, skipping non-data rows."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb["Data Entry"]

    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        product_type = row[0]
        if (
            not product_type
            or str(product_type).strip().upper() not in VALID_PRODUCT_TYPES
        ):
            continue
        cleaned = {}
        for key, val in zip(headers, row):
            if isinstance(val, str) and not val.strip().replace("\xa0", ""):
                val = None
            cleaned[key] = val
        rows.append(cleaned)

    return rows


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("ES Windows — Order Form Automation")
        self.root.resizable(False, False)

        self.driver = None
        self.excel_path = None
        self._stop_requested = False

        self._build_ui()

    # ─── UI Layout ──────────────────────────────────────────

    def _build_ui(self):
        pad = {"padx": 12, "pady": 6}

        # ── Step 1: Browser ──────────────────────────────────
        step1 = tk.LabelFrame(
            self.root, text="Step 1 — Browser", font=("Helvetica", 13, "bold")
        )
        step1.pack(fill="x", **pad)

        self.browser_status = tk.Label(
            step1, text="Not connected", fg="red", anchor="w"
        )
        self.browser_status.pack(side="left", padx=8, pady=8, fill="x", expand=True)

        self.btn_open_browser = tk.Button(
            step1, text="Open Browser", width=14, command=self._open_browser
        )
        self.btn_open_browser.pack(side="right", padx=8, pady=8)

        # ── Step 2: Order details ────────────────────────────
        step2 = tk.LabelFrame(
            self.root, text="Step 2 — Order Details", font=("Helvetica", 13, "bold")
        )
        step2.pack(fill="x", **pad)

        # Order number
        row_order = tk.Frame(step2)
        row_order.pack(fill="x", padx=8, pady=(8, 4))
        tk.Label(row_order, text="Order Number:", width=14, anchor="w").pack(
            side="left"
        )
        self.entry_order = tk.Entry(row_order, width=30)
        self.entry_order.pack(side="left", padx=(4, 0))

        # Excel file
        row_file = tk.Frame(step2)
        row_file.pack(fill="x", padx=8, pady=(4, 4))
        tk.Label(row_file, text="Excel File:", width=14, anchor="w").pack(side="left")
        self.lbl_file = tk.Label(
            row_file, text="No file selected", fg="gray", anchor="w"
        )
        self.lbl_file.pack(side="left", padx=(4, 8), fill="x", expand=True)
        tk.Button(row_file, text="Browse...", command=self._browse_file).pack(
            side="right"
        )

        # Start from row
        row_start = tk.Frame(step2)
        row_start.pack(fill="x", padx=8, pady=(4, 8))
        tk.Label(row_start, text="Start from row:", width=14, anchor="w").pack(
            side="left"
        )
        self.entry_start_row = tk.Entry(row_start, width=8)
        self.entry_start_row.insert(0, "1")
        self.entry_start_row.pack(side="left", padx=(4, 0))
        tk.Label(
            row_start, text="(use this to resume after a failure)", fg="gray"
        ).pack(side="left", padx=(8, 0))

        # ── Step 3: Run ──────────────────────────────────────
        step3 = tk.LabelFrame(
            self.root, text="Step 3 — Run", font=("Helvetica", 13, "bold")
        )
        step3.pack(fill="x", **pad)

        self.btn_start = tk.Button(
            step3,
            text="Start Automation",
            font=("Helvetica", 12, "bold"),
            bg="#4CAF50",
            fg="white",
            height=2,
            command=self._start,
        )
        self.btn_start.pack(fill="x", padx=8, pady=(8, 4))

        # Stop button (hidden until automation starts)
        self.controls_frame = tk.Frame(step3)
        self.controls_frame.pack(fill="x", padx=8, pady=(0, 8))

        style = ttk.Style()
        style.configure(
            "Stop.TButton",
            font=("Helvetica", 11, "bold"),
            foreground="#F44336",
        )

        self.btn_stop = ttk.Button(
            self.controls_frame,
            text="⏹  Stop",
            style="Stop.TButton",
            command=self._stop,
        )

        # ── Log area ─────────────────────────────────────────
        log_frame = tk.LabelFrame(
            self.root, text="Progress", font=("Helvetica", 13, "bold")
        )
        log_frame.pack(fill="both", expand=True, **pad)

        self.log_text = tk.Text(
            log_frame,
            height=16,
            width=72,
            state="disabled",
            wrap="word",
            font=("Courier", 11),
        )
        scrollbar = tk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True, padx=4, pady=4)

    # ─── Logging ────────────────────────────────────────────

    def _log(self, message):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    # ─── Step 1: Open Browser ───────────────────────────────

    def _open_browser(self):
        if self.driver:
            messagebox.showinfo("Browser", "Browser is already open.")
            return

        self.btn_open_browser.configure(state="disabled", text="Opening...")
        self._log("Launching Chrome...")

        threading.Thread(target=self._launch_chrome, daemon=True).start()

    def _launch_chrome(self):
        try:
            options = Options()
            options.add_argument(f"--user-data-dir={CHROME_PROFILE_DIR}")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])

            from webdriver_manager.chrome import ChromeDriverManager
            self.root.after(0, lambda: self._log("Downloading ChromeDriver (first time only)..."))
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)

            self.root.after(0, self._on_browser_ready)
        except Exception as e:
            self.root.after(0, lambda: self._on_browser_error(str(e)))

    def _on_browser_ready(self):
        self.browser_status.configure(
            text="Connected — log in, then continue here", fg="green"
        )
        self.btn_open_browser.configure(state="normal", text="Open Browser")
        self._log("Chrome opened. Please navigate to orders.eswindows.co,")
        self._log("log in, and then come back here to proceed.\n")

    def _on_browser_error(self, error):
        self.browser_status.configure(text="Failed to connect", fg="red")
        self.btn_open_browser.configure(state="normal", text="Open Browser")
        self._log(f"ERROR: Could not launch Chrome: {error}")
        self._log("Make sure Google Chrome and ChromeDriver are installed.")

    # ─── Stop Control ───────────────────────────────────────

    def _show_controls(self):
        """Show stop button during automation."""
        self.btn_stop.pack(fill="x")

    def _hide_controls(self):
        """Hide stop button when automation is not running."""
        self.btn_stop.pack_forget()

    def _stop(self):
        self._stop_requested = True
        self._log("  >> STOPPING after current line item...")

    # ─── Step 2: Browse file ────────────────────────────────

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if path:
            self.excel_path = path
            filename = os.path.basename(path)
            self.lbl_file.configure(text=filename, fg="black")

    # ─── Step 3: Start Automation ───────────────────────────

    def _start(self):
        # Validate inputs
        if not self.driver:
            messagebox.showwarning(
                "Browser Required", "Please open the browser first (Step 1)."
            )
            return

        order_number = self.entry_order.get().strip()
        if not order_number:
            messagebox.showwarning("Missing Order", "Please enter an order number.")
            return

        if not self.excel_path:
            messagebox.showwarning("Missing File", "Please select an Excel file.")
            return

        # Validate "Start from row"
        start_row_raw = self.entry_start_row.get().strip()
        try:
            start_row = int(start_row_raw)
        except ValueError:
            messagebox.showwarning(
                "Invalid Start Row",
                f"'Start from row' must be a number. Got: '{start_row_raw}'",
            )
            return
        if start_row <= 0:
            messagebox.showwarning(
                "Invalid Start Row", "'Start from row' must be 1 or greater."
            )
            return

        # Read Excel (re-read from disk so edits are picked up)
        try:
            rows = read_excel(self.excel_path)
        except FileNotFoundError:
            messagebox.showerror("File Error", f"File not found:\n{self.excel_path}")
            return
        except Exception as e:
            messagebox.showerror("File Error", f"Could not read Excel file:\n{e}")
            return

        if not rows:
            messagebox.showinfo(
                "No Data", "No valid line items found in the 'Data Entry' sheet."
            )
            return

        if start_row > len(rows):
            messagebox.showwarning(
                "Invalid Start Row",
                f"'Start from row' ({start_row}) is greater than the total line items ({len(rows)}).",
            )
            return

        self._log(f"Found {len(rows)} line items.")
        if start_row > 1:
            self._log(
                f"Resuming from row {start_row} ({len(rows) - start_row + 1} to process)."
            )
        for i, row in enumerate(rows[:5]):
            self._log(
                f"  {i+1}. {row.get('Product Type', '?')} | "
                f"{row.get('Model', '?')} | "
                f"W:{row.get('Width', '?')} x H:{row.get('Height', '?')}"
            )
        if len(rows) > 5:
            self._log(f"  ... and {len(rows) - 5} more")

        prompt = (
            f"Fill rows {start_row}–{len(rows)} into order #{order_number}?"
            if start_row > 1
            else f"Fill {len(rows)} line items into order #{order_number}?"
        )
        confirm = messagebox.askyesno("Confirm", prompt)
        if not confirm:
            self._log("Cancelled.\n")
            return

        # Disable UI during automation and show stop button
        self._stop_requested = False
        self.btn_start.configure(state="disabled", text="Running...")
        self.btn_open_browser.configure(state="disabled")
        self._show_controls()

        threading.Thread(
            target=self._run_automation,
            args=(order_number, rows, start_row),
            daemon=True,
        ).start()

    def _run_automation(self, order_number, rows, start_row=1):
        try:
            url = sel.ORDER_URL.format(order_number=order_number)
            self.root.after(
                0, lambda: self._log(f"\nNavigating to order #{order_number}...")
            )
            self.driver.get(url)
            time.sleep(4)

            filler = FormFiller(self.driver)
            success_count = 0
            total = len(rows)

            # Slice rows from start_row (1-indexed)
            rows_to_process = rows[start_row - 1 :]

            for idx, row in enumerate(rows_to_process):
                # ── Check for stop ──
                if self._stop_requested:
                    row_num = start_row + idx
                    self.root.after(
                        0,
                        lambda rn=row_num, sc=success_count: self._log(
                            f"\n  >> STOPPED by user at line item {rn}.\n"
                            f"  Successfully created {sc} line items in this run.\n"
                            f"  To resume: set 'Start from row' to {rn} and run again."
                        ),
                    )
                    self.root.after(0, self._on_automation_done)
                    return

                row_num = start_row + idx  # original row number (1-indexed)

                self.root.after(
                    0,
                    lambda rn=row_num, row=row: self._log(
                        f"\n{'─' * 45}\n"
                        f"Line Item {rn}/{total}\n"
                        f"  Product: {row.get('Product Type', '?')}\n"
                        f"  Model:   {row.get('Model', '?')}\n"
                        f"  Size:    {row.get('Width', '?')} x {row.get('Height', '?')}"
                    ),
                )

                try:
                    filler.add_line_item(row)
                    success_count += 1
                    self.root.after(0, lambda: self._log("  >> CREATED successfully"))
                    time.sleep(3)
                except FormFillerError as e:
                    self.root.after(
                        0,
                        lambda e=e, rn=row_num, sc=success_count: self._log(
                            f"\n  >> {e}\n"
                            f"  Process terminated at line item {rn}.\n"
                            f"  Successfully created {sc} line items in this run.\n"
                            f"  To resume: fix the Excel, set 'Start from row' to {rn}, and run again."
                        ),
                    )
                    self.root.after(0, self._on_automation_done)
                    return
                except Exception as e:
                    self.root.after(
                        0,
                        lambda e=e, rn=row_num, sc=success_count: self._log(
                            f"\n  >> UNEXPECTED ERROR: {e}\n"
                            f"  Process terminated at line item {rn}.\n"
                            f"  Successfully created {sc} line items in this run.\n"
                            f"  To resume: fix the issue, set 'Start from row' to {rn}, and run again."
                        ),
                    )
                    self.root.after(0, self._on_automation_done)
                    return

            self.root.after(
                0,
                lambda sc=success_count: self._log(
                    f"\n{'=' * 45}\n"
                    f"DONE! Successfully created {sc} line items in this run.\n"
                    f"{'=' * 45}"
                ),
            )

        except Exception as e:
            self.root.after(0, lambda e=e: self._log(f"\nERROR: {e}"))

        self.root.after(0, self._on_automation_done)

    def _on_automation_done(self):
        self.btn_start.configure(state="normal", text="Start Automation")
        self.btn_open_browser.configure(state="normal")
        self._hide_controls()


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
